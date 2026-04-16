import re
import pandas as pd
from openpyxl import load_workbook

# ── 欄位名稱對照表（各廠來源欄位 → 統一欄位名稱） ──────────────────────────
COLUMN_MAP = {
    # 營收類
    "Net Operating Revenues": "revenue",
    "營業收入": "revenue",
    "revenue":  "revenue",
    "net revenue": "revenue",
    "sales":    "revenue",
    # COGS 類
    "銷售成本": "cogs",
    "製造成本": "cogs",
    "cogs":     "cogs",
    "Cost of goods sold": "cogs",
    "cost of sales": "cogs",
    # OpEx 類
    "營業費用": "opex",
    "operating expense": "opex",
    "opex":     "opex",
    "operating expenses": "opex",
}

# ── Auto-ID：從 Excel 抓取廠別 / 年度 / 季度 ──────────────────────────────
_QUARTER_RE = re.compile(r"Q[1-4]", re.IGNORECASE)
_YEAR_RE    = re.compile(r"20\d{2}")

_PROBE_CELLS = ["A1", "B1", "C1", "A2", "B2", "C2",
                "A3", "B3", "C3", "D1", "E1"]


def auto_identify(file) -> dict:
    """
    讀取 Excel 前幾個儲存格，嘗試辨識：
      - site_id   (廠別)
      - year      (年度)
      - quarter   (季度 Q1~Q4)
      - data_type (Actual / Budget)
    回傳 dict，辨識失敗的欄位值為 None。
    """
    wb = load_workbook(file, read_only=True, data_only=True)
    ws = wb.active

    site_id   = None
    year      = None
    quarter   = None
    data_type = "Actual"

    # 收集前幾個非空的儲存格文字
    cell_texts = []
    for cell_addr in _PROBE_CELLS:
        try:
            val = ws[cell_addr].value
            if val is not None:
                cell_texts.append(str(val).strip())
        except Exception:
            pass

    # 也掃第一行所有欄
    for cell in next(ws.iter_rows(max_row=1), []):
        if cell.value:
            cell_texts.append(str(cell.value).strip())

    combined = " ".join(cell_texts)

    # 年度
    y_match = _YEAR_RE.search(combined)
    if y_match:
        year = y_match.group()

    # 季度
    q_match = _QUARTER_RE.search(combined)
    if q_match:
        quarter = q_match.group().upper()

    # 廠別：取第一個看起來像 Site 名的詞
    site_match = re.search(r"(Site\s*[A-Z0-9]+|廠[別區]?\s*[A-Z0-9]+|[A-Z]{2,}廠)", combined, re.IGNORECASE)
    if site_match:
        site_id = site_match.group().strip()
    else:
        # fallback：取第一個非空儲存格
        site_id = cell_texts[0] if cell_texts else None

    # data_type
    if re.search(r"budget|預算", combined, re.IGNORECASE):
        data_type = "Budget"

    wb.close()

    # 組合 year_month 格式：2025-Q1
    year_month = None
    if year and quarter:
        year_month = f"{year}-{quarter}"

    return {
        "site_id":    site_id,
        "year":       year,
        "quarter":    quarter,
        "year_month": year_month,
        "data_type":  data_type,
        "raw_texts":  cell_texts[:10],
    }


# ── 數字標準化 ────────────────────────────────────────────────────────────────
def _to_float(val) -> float:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace(",", "").replace(" ", "").replace("$", "").replace("¥", "")
    try:
        return float(s)
    except ValueError:
        return 0.0


# ── 主要 ETL：DataFrame → 標準化 records ────────────────────────────────────
def transform(df: pd.DataFrame, meta: dict) -> list[dict]:
    """
    將上傳的 DataFrame 清洗成可寫入 DB 的 records 列表。
    meta 須含 site_id, year_month, data_type。
    """
    # 統一小寫欄位名稱
    df.columns = [str(c).strip().lower() for c in df.columns]

    # 套用欄位對照表
    rename = {k.lower(): v for k, v in COLUMN_MAP.items()}
    df = df.rename(columns=rename)

    # 確保三欄都存在
    for col in ("revenue", "cogs", "opex"):
        if col not in df.columns:
            df[col] = 0.0

    records = []
    for _, row in df.iterrows():
        rev  = _to_float(row.get("revenue"))
        cogs = _to_float(row.get("cogs"))
        opex = _to_float(row.get("opex"))

        # 過濾全零行
        if rev == 0 and cogs == 0 and opex == 0:
            continue

        records.append({
            "site_id":    meta["site_id"],
            "year_month": meta["year_month"],
            "data_type":  meta["data_type"],
            "revenue":    rev,
            "cogs":       cogs,
            "opex":       opex,
        })

    return records
